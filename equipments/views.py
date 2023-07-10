from .models import Equipment
from .serializers import EquipmentSerializer
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import landscape, A4, letter
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from xhtml2pdf import pisa, default
from xhtml2pdf.default import DEFAULT_FONT
from rest_framework.generics import GenericAPIView
from rest_framework.mixins import ListModelMixin, CreateModelMixin, UpdateModelMixin, DestroyModelMixin, RetrieveModelMixin

class EquipmentLC(GenericAPIView, ListModelMixin, CreateModelMixin):
    queryset = Equipment.objects.all()                   #Retrieved all instances from database
    serializer_class = EquipmentSerializer               #Serialize or deserialize the data

    def get(self, request, *args, **kwargs):             #Get all equipments
        return self.list(request, *args, **kwargs)
    def post(self,request,*args,  **kwargs):             #Create new equipment
        return self.create(request, *args, **kwargs)

class EquipmentRUD(GenericAPIView, RetrieveModelMixin, UpdateModelMixin, DestroyModelMixin):
    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer

    def get(self, request, *args, **kwargs):             #Get one equipment
        return self.retrieve(request, *args, **kwargs)
    def put(self,request,*args,  **kwargs):              #Update one equipment
        return self.update(request, *args, **kwargs)
    def delete(self,request,*args,  **kwargs):           #Delete one equipment
        return self.destroy(request, *args, **kwargs)

def export_to_excel(request):
    # Generate Excel file
    workbook = Workbook()
    worksheet = workbook.active

    # Get all Equipment instances
    queryset = Equipment.objects.all()

    # Write headers
    headers = ['Bil', 'No. Siri Pendaftaran', 'Name Aset', 'Lokasi', 'Kuantiti', 'Keadaan Harta Tetap', 'Daftar KEW.PA', 'Harga Aset (RM)', 'Catatan']
    worksheet.append(headers)

    # Write data rows
    data_rows = []
    for index, equipment in enumerate(queryset, start=1):
        if equipment.registered:
            registered = '/'
        else:
            registered = ''

        price = round(equipment.price, 2)
        row = [index, equipment.regNum, equipment.name, equipment.location, equipment.quantity, equipment.status, registered, price, '']
        data_rows.append(row)
        worksheet.append(row)

    # Set column widths
    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        data_width = max(len(str(row[column_index - 1])) for row in data_rows) + 2
        header_width = len(header) + 2
        column_width = max(data_width, header_width)
        worksheet.column_dimensions[column_letter].width = column_width
        
    # Set response headers for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def export_to_pdf(request):
    # Get all Equipment instances
    queryset = Equipment.objects.all()

    # Define the HTML content
    html_content = f"""
    <html>
    <head>
        <style>
            
            body {{
                font-family: Arial, sans-serif;
                font-size: 12px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
            }}
            th, td {{
                border: 1px solid black;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
            header {{
                font-size: 16px;
                font-weight: bold;
                margin-bottom: 10px;
                text-align: center;
            }}
        </style>
    </head>
    <body>
            <div><header>LAPORAN PEMERIKSAAN INVENTORI</header></div>
            <div><p>Tahun: ________________</p>
            <p>Unit/Makmal: ______________</p></div>
            <p>Fakulti/PTJ: ______________</p>
            
        <table>
            <thead>
                <tr>
                    <th>Bil</th>
                    <th>No. Siri Pendaftaran</th>
                    <th>Name Aset</th>
                    <th>Lokasi</th>
                    <th>Kuantiti</th>
                    <th>Keadaan Harta Tetap</th>
                    <th>Harga Aset</th>
                </tr>
            </thead>
            <tbody>
    """

    # Add data rows to the HTML content
    for i, equipment in enumerate(queryset):
        row = f"""
            <tr>
                <td>{i + 1}</td>
                <td>{equipment.regNum}</td>
                <td>{equipment.name}</td>
                <td>{equipment.location}</td>
                <td>{equipment.quantity}</td>
                <td>{equipment.status}</td>
                <td>{round(equipment.price, 2)}</td>
            </tr>
        """
        html_content += row

    # Close the HTML content
    html_content += """
            </tbody>
        </table>
        <div>
            <div><p>________________</p><p>(Tandatangan)</p></div>
            <div><p>________________</p><p>(Nama Pegawai Pemeriksa)</p></div>
        </div>
        <div>
            <div><p>________________</p><p>(Jawatan)</p></div>
            <div><p>________________</p><p>(Tarikh)</p></div>
        </div>
    </body>
    </html>
    """

    # Create a file-like buffer to receive PDF data
    pdf_buffer = BytesIO()
    pisa.showLogging()

    # Create the PDF using the HTML content
    pisaStatus = pisa.CreatePDF(html_content, dest=pdf_buffer, orientation='Landscape')

    # Check if PDF generation was successful
    if pisaStatus.err:
        return HttpResponse('PDF generation failed.', status=500)

    # Set the appropriate PDF response headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="data.pdf"'

    # Get the PDF content from the buffer and write it to the response
    pdf_buffer.seek(0)
    response.write(pdf_buffer.read())

    return response